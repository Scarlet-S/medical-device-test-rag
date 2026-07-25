import argparse
import csv
import json
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
BASELINE_PATH = (
    PROJECT_ROOT
    / "evaluation"
    / "baseline"
    / "医疗器械软件测试知识库_评测工作簿_v1.xlsx"
)
RESULTS_DIR = PROJECT_ROOT / "evaluation" / "results"
SHEET_NAME = "评测记录"
HEADER_ROW = 2


def load_manual_baseline():
    workbook = load_workbook(
        BASELINE_PATH,
        read_only=True,
        data_only=True,
    )
    sheet = workbook[SHEET_NAME]

    headers = {
        str(cell.value).strip(): cell.column
        for cell in sheet[HEADER_ROW]
        if cell.value
    }

    required = {
        "问题ID",
        "预期文档",
        "实际Top1文档",
        "Top1命中",
        "Top3命中",
        "引用正确",
        "回答准确度(0-2)",
        "是否幻觉",
        "综合得分",
    }

    missing = required - set(headers)
    if missing:
        workbook.close()
        raise RuntimeError(
            f"人工基线缺少列：{', '.join(sorted(missing))}"
        )

    records = {}

    for row in range(HEADER_ROW + 1, sheet.max_row + 1):
        question_id = sheet.cell(
            row,
            headers["问题ID"],
        ).value

        if not question_id:
            continue

        question_id = str(question_id).strip()

        records[question_id] = {
            "question_id": question_id,
            "expected_document": str(
                sheet.cell(
                    row,
                    headers["预期文档"],
                ).value
                or ""
            ).strip(),
            "manual_actual_top1": str(
                sheet.cell(
                    row,
                    headers["实际Top1文档"],
                ).value
                or ""
            ).strip(),
            "manual_top1": int(
                sheet.cell(
                    row,
                    headers["Top1命中"],
                ).value
                or 0
            ),
            "manual_top3": int(
                sheet.cell(
                    row,
                    headers["Top3命中"],
                ).value
                or 0
            ),
            "manual_citation_correct": int(
                sheet.cell(
                    row,
                    headers["引用正确"],
                ).value
                or 0
            ),
            "manual_answer_accuracy": int(
                sheet.cell(
                    row,
                    headers["回答准确度(0-2)"],
                ).value
                or 0
            ),
            "manual_hallucination": int(
                sheet.cell(
                    row,
                    headers["是否幻觉"],
                ).value
                or 0
            ),
            "manual_score": float(
                sheet.cell(
                    row,
                    headers["综合得分"],
                ).value
                or 0
            ),
        }

    workbook.close()
    return records


def find_latest_result():
    candidates = sorted(
        RESULTS_DIR.glob("batch_eval_50_*.json"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )

    if not candidates:
        raise RuntimeError("未找到50题自动评测JSON结果")

    return candidates[0]


def load_auto_results(path):
    with path.open("r", encoding="utf-8") as file:
        payload = json.load(file)

    return {
        record["question_id"]: record
        for record in payload["results"]
    }


def change_label(manual_value, auto_value):
    if manual_value == auto_value:
        return "一致"

    if manual_value == 1 and auto_value == 0:
        return "本轮下降"

    return "本轮提升"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--result",
        type=Path,
        help="自动评测JSON路径；不填则读取最新50题结果",
    )
    args = parser.parse_args()

    try:
        result_path = (
            args.result.resolve()
            if args.result
            else find_latest_result()
        )
        manual = load_manual_baseline()
        automatic = load_auto_results(result_path)
    except Exception as exc:
        print(f"比较失败：{exc}")
        return 1

    comparisons = []

    for question_id, manual_record in manual.items():
        auto_record = automatic.get(question_id)

        if not auto_record:
            continue

        auto_top1 = int(auto_record["top1_hit"])
        auto_top3 = int(auto_record["top3_hit"])

        comparisons.append(
            {
                **manual_record,
                "auto_actual_top1": auto_record[
                    "actual_top1_document"
                ],
                "auto_top1": auto_top1,
                "auto_top3": auto_top3,
                "top1_change": change_label(
                    manual_record["manual_top1"],
                    auto_top1,
                ),
                "top3_change": change_label(
                    manual_record["manual_top3"],
                    auto_top3,
                ),
                "auto_attempts": auto_record["attempts"],
                "auto_reference_count": auto_record[
                    "reference_count"
                ],
            }
        )

    if not comparisons:
        print("比较失败：没有匹配的问题ID")
        return 1

    manual_top1 = sum(
        row["manual_top1"] for row in comparisons
    )
    auto_top1 = sum(
        row["auto_top1"] for row in comparisons
    )
    manual_top3 = sum(
        row["manual_top3"] for row in comparisons
    )
    auto_top3 = sum(
        row["auto_top3"] for row in comparisons
    )

    top1_regressions = [
        row["question_id"]
        for row in comparisons
        if row["top1_change"] == "本轮下降"
    ]
    top1_improvements = [
        row["question_id"]
        for row in comparisons
        if row["top1_change"] == "本轮提升"
    ]
    top3_regressions = [
        row["question_id"]
        for row in comparisons
        if row["top3_change"] == "本轮下降"
    ]

    output_path = (
        RESULTS_DIR
        / f"comparison_{result_path.stem}.csv"
    )

    fieldnames = list(comparisons[0].keys())

    with output_path.open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as file:
        writer = csv.DictWriter(
            file,
            fieldnames=fieldnames,
        )
        writer.writeheader()
        writer.writerows(comparisons)

    total = len(comparisons)

    print("=" * 60)
    print(f"比较题数：{total}")
    print(
        f"人工Top1：{manual_top1}/{total}"
        f"（{manual_top1 / total:.1%}）"
    )
    print(
        f"自动Top1：{auto_top1}/{total}"
        f"（{auto_top1 / total:.1%}）"
    )
    print(
        f"人工Top3：{manual_top3}/{total}"
        f"（{manual_top3 / total:.1%}）"
    )
    print(
        f"自动Top3：{auto_top3}/{total}"
        f"（{auto_top3 / total:.1%}）"
    )
    print(
        "Top1本轮下降："
        + (
            "、".join(top1_regressions)
            if top1_regressions
            else "无"
        )
    )
    print(
        "Top1本轮提升："
        + (
            "、".join(top1_improvements)
            if top1_improvements
            else "无"
        )
    )
    print(
        "Top3本轮下降："
        + (
            "、".join(top3_regressions)
            if top3_regressions
            else "无"
        )
    )
    print(f"对比结果：{output_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())