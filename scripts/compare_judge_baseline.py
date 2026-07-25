import csv
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
BASELINE_DIR = PROJECT_ROOT / "evaluation" / "baseline"
RESULTS_DIR = PROJECT_ROOT / "evaluation" / "results"
SHEET_NAME = "评测记录"
HEADER_ROW = 2


def find_latest_judge_result():
    candidates = sorted(
        RESULTS_DIR.glob("judge_eval_50_*.json"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )

    if not candidates:
        raise RuntimeError("未找到50题自动裁判JSON结果")

    return candidates[0]


def find_baseline_workbook():
    candidates = sorted(BASELINE_DIR.glob("*.xlsx"))

    if not candidates:
        raise RuntimeError("未找到人工基线工作簿")

    return candidates[0]


def to_int(value):
    if value is None or value == "":
        return None

    return int(value)


def load_human_scores(path):
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )
    sheet = workbook[SHEET_NAME]

    headers = {
        str(cell.value).strip(): cell.column
        for cell in sheet[HEADER_ROW]
        if cell.value
    }
    required = {
        "问题ID",
        "引用正确",
        "回答准确度(0-2)",
        "是否幻觉",
    }
    missing = required - set(headers)

    if missing:
        workbook.close()
        raise RuntimeError(
            f"人工基线缺少列：{', '.join(sorted(missing))}"
        )

    scores = {}

    for row in range(HEADER_ROW + 1, sheet.max_row + 1):
        question_id = sheet.cell(
            row,
            headers["问题ID"],
        ).value

        if not question_id:
            continue

        scores[str(question_id).strip()] = {
            "human_citation_correct": to_int(
                sheet.cell(
                    row,
                    headers["引用正确"],
                ).value
            ),
            "human_answer_accuracy": to_int(
                sheet.cell(
                    row,
                    headers["回答准确度(0-2)"],
                ).value
            ),
            "human_hallucination": to_int(
                sheet.cell(
                    row,
                    headers["是否幻觉"],
                ).value
            ),
        }

    workbook.close()
    return scores


def load_judge_scores(path):
    with path.open("r", encoding="utf-8") as file:
        payload = json.load(file)

    return {
        record["question_id"]: record
        for record in payload.get("results", [])
        if record.get("status") == "success"
    }


def is_match(human_value, judge_value):
    if human_value is None:
        return None

    return int(human_value == judge_value)


def compare_scores(human_scores, judge_scores):
    rows = []

    for question_id in sorted(human_scores):
        if question_id not in judge_scores:
            continue

        human = human_scores[question_id]
        judge = judge_scores[question_id]

        citation_match = is_match(
            human["human_citation_correct"],
            judge["citation_correct"],
        )
        accuracy_match = is_match(
            human["human_answer_accuracy"],
            judge["answer_accuracy"],
        )
        hallucination_match = is_match(
            human["human_hallucination"],
            judge["hallucination"],
        )
        comparable_matches = [
            value
            for value in (
                citation_match,
                accuracy_match,
                hallucination_match,
            )
            if value is not None
        ]

        rows.append(
            {
                "question_id": question_id,
                "question": judge.get("question", ""),
                **human,
                "judge_citation_correct": judge[
                    "citation_correct"
                ],
                "citation_match": citation_match,
                "judge_answer_accuracy": judge[
                    "answer_accuracy"
                ],
                "accuracy_match": accuracy_match,
                "judge_hallucination": judge[
                    "hallucination"
                ],
                "hallucination_match": hallucination_match,
                "all_available_match": int(
                    bool(comparable_matches)
                    and all(comparable_matches)
                ),
                "judge_reason": judge.get("reason", ""),
            }
        )

    return rows


def rate(numerator, denominator):
    return (
        f"{numerator}/{denominator}"
        f"（{numerator / denominator:.1%}）"
        if denominator
        else "无可比较数据"
    )


def summarize(rows, match_field):
    values = [
        row[match_field]
        for row in rows
        if row[match_field] is not None
    ]
    return sum(values), len(values)


def save_csv(rows, judge_path):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = (
        RESULTS_DIR
        / f"judge_baseline_comparison_{timestamp}.csv"
    )
    fieldnames = list(rows[0].keys())

    with output_path.open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    return output_path


def main():
    try:
        judge_path = find_latest_judge_result()
        workbook_path = find_baseline_workbook()
        human_scores = load_human_scores(workbook_path)
        judge_scores = load_judge_scores(judge_path)
        rows = compare_scores(human_scores, judge_scores)
    except Exception as exc:
        print(f"比较失败：{exc}")
        return 1

    if not rows:
        print("比较失败：没有可对齐的题目")
        return 1

    citation_hits, citation_total = summarize(
        rows,
        "citation_match",
    )
    accuracy_hits, accuracy_total = summarize(
        rows,
        "accuracy_match",
    )
    hallucination_hits, hallucination_total = summarize(
        rows,
        "hallucination_match",
    )
    full_hits, full_total = summarize(
        rows,
        "all_available_match",
    )

    citation_diff = [
        row["question_id"]
        for row in rows
        if row["citation_match"] == 0
    ]
    accuracy_diff = [
        row["question_id"]
        for row in rows
        if row["accuracy_match"] == 0
    ]
    hallucination_diff = [
        row["question_id"]
        for row in rows
        if row["hallucination_match"] == 0
    ]

    output_path = save_csv(rows, judge_path)

    print("=" * 60)
    print(f"比较题数：{len(rows)}")
    print(
        "引用正确一致："
        + rate(citation_hits, citation_total)
    )
    print(
        "回答准确度一致："
        + rate(accuracy_hits, accuracy_total)
    )
    print(
        "幻觉判断一致："
        + rate(hallucination_hits, hallucination_total)
    )
    print(
        "三项全部一致："
        + rate(full_hits, full_total)
    )
    print(
        "引用正确争议："
        + ("、".join(citation_diff) if citation_diff else "无")
    )
    print(
        "回答准确度争议："
        + ("、".join(accuracy_diff) if accuracy_diff else "无")
    )
    print(
        "幻觉判断争议："
        + (
            "、".join(hallucination_diff)
            if hallucination_diff
            else "无"
        )
    )
    print(f"人工基线：{workbook_path}")
    print(f"自动裁判：{judge_path}")
    print(f"对比结果：{output_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
