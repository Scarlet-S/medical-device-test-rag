import argparse
import csv
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from ragflow_client import RAGFlowClient


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK_PATH = (
    PROJECT_ROOT
    / "evaluation"
    / "baseline"
    / "医疗器械软件测试知识库_评测工作簿_v1.xlsx"
)
RESULTS_DIR = PROJECT_ROOT / "evaluation" / "results"
SHEET_NAME = "评测题集"
HEADER_ROW = 2
MAX_ATTEMPTS = 3


def read_json_test_cases(
    limit,
    cases_path,
    question_id_filter="",
    question_id_filters=None,
):
    with cases_path.open("r", encoding="utf-8") as file:
        payload = json.load(file)

    raw_cases = payload.get("cases", [])
    if not isinstance(raw_cases, list):
        raise RuntimeError("JSON题集的cases字段必须是数组")

    required_fields = {
        "question_id",
        "expected_document",
        "question",
        "expected_location",
        "reference_answer",
    }
    cases = []
    filters = {
        value.strip().upper()
        for value in (question_id_filters or [])
        if value.strip()
    }
    if question_id_filter:
        filters.add(question_id_filter)

    for raw_case in raw_cases:
        if not isinstance(raw_case, dict):
            raise RuntimeError("JSON题集中的每道题必须是对象")

        missing = [
            field
            for field in required_fields
            if not str(raw_case.get(field, "")).strip()
        ]
        if missing:
            raise RuntimeError(
                "JSON题集存在缺失字段："
                + ", ".join(sorted(missing))
            )

        question_id = str(raw_case["question_id"]).strip()
        if (
            filters
            and question_id.upper() not in filters
        ):
            continue

        cases.append(
            {
                "question_id": question_id,
                "expected_document": str(
                    raw_case["expected_document"]
                ).strip(),
                "expected_filename": str(
                    raw_case.get("expected_filename", "")
                ).strip(),
                "question": str(raw_case["question"]).strip(),
                "expected_location": str(
                    raw_case["expected_location"]
                ).strip(),
                "reference_answer": str(
                    raw_case["reference_answer"]
                ).strip(),
            }
        )

        if len(cases) >= limit:
            break

    return cases


def extract_doc_code(value):
    if not value:
        return ""

    match = re.search(
        r"(?:FDAAI_K\d{6}|PRACTICE\d{3}|DOC\d{3})",
        str(value),
        re.IGNORECASE,
    )
    return match.group(0).upper() if match else ""


def read_test_cases(
    limit,
    workbook_path,
    question_id_filter="",
    question_id_filters=None,
):
    if workbook_path.suffix.casefold() == ".json":
        return read_json_test_cases(
            limit,
            workbook_path,
            question_id_filter,
            question_id_filters,
        )

    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )
    sheet = workbook[SHEET_NAME]

    # Some valid XLSX generators omit worksheet dimension metadata.
    # Force openpyxl to calculate it before using max_row/max_column.
    if sheet.max_row is None or sheet.max_column is None:
        sheet.calculate_dimension(force=True)

    headers = {
        str(cell.value).strip(): cell.column
        for cell in sheet[HEADER_ROW]
        if cell.value
    }

    required_headers = {
        "问题ID",
        "预期文档",
        "评测问题",
        "预期定位",
        "标准答案要点",
    }

    missing = required_headers - set(headers)
    if missing:
        workbook.close()
        raise RuntimeError(
            f"评测题集缺少列：{', '.join(sorted(missing))}"
        )

    cases = []
    filters = {
        value.strip().upper()
        for value in (question_id_filters or [])
        if value.strip()
    }
    if question_id_filter:
        filters.add(question_id_filter)

    for row_number in range(HEADER_ROW + 1, sheet.max_row + 1):
        question_id = sheet.cell(
            row_number,
            headers["问题ID"],
        ).value

        question = sheet.cell(
            row_number,
            headers["评测问题"],
        ).value

        if not question_id or not question:
            continue

        normalized_question_id = str(question_id).strip()
        if (
            filters
            and normalized_question_id.upper()
            not in filters
        ):
            continue

        cases.append(
            {
                "question_id": normalized_question_id,
                "expected_document": str(
                    sheet.cell(
                        row_number,
                        headers["预期文档"],
                    ).value
                    or ""
                ).strip(),
                "expected_filename": "",
                "question": str(question).strip(),
                "expected_location": str(
                    sheet.cell(
                        row_number,
                        headers["预期定位"],
                    ).value
                    or ""
                ).strip(),
                "reference_answer": str(
                    sheet.cell(
                        row_number,
                        headers["标准答案要点"],
                    ).value
                    or ""
                ).strip(),
            }
        )

        if len(cases) >= limit:
            break

    workbook.close()
    return cases


def compact_reference(chunk):
    document_name = (
        chunk.get("document_name")
        or chunk.get("doc_name")
        or ""
    )

    return {
        "document_name": document_name,
        "document_code": extract_doc_code(document_name),
        "similarity": chunk.get("similarity"),
        "chunk_id": chunk.get("id", ""),
        "content": chunk.get("content", ""),
        "positions": chunk.get("positions", []),
    }


def evaluate_case(client, case):
    started = time.perf_counter()
    result = None
    error = ""
    attempts = 0

    for attempts in range(1, MAX_ATTEMPTS + 1):
        try:
            candidate = client.ask(case["question"])

            if candidate["answer"].strip():
                result = candidate
                break

            error = "RAGFlow返回空答案"
        except Exception as exc:
            error = f"{type(exc).__name__}: {exc}"

        if attempts < MAX_ATTEMPTS:
            print(
                f"  第{attempts}次未成功，"
                "2秒后重试……"
            )
            time.sleep(2)

    elapsed = round(time.perf_counter() - started, 2)

    references = []
    answer = ""
    session_id = ""

    if result:
        answer = result["answer"]
        session_id = result.get("session_id", "")
        references = [
            compact_reference(chunk)
            for chunk in result["references"]
        ]

    expected_code = extract_doc_code(
        case["expected_document"]
    )
    retrieved_codes = [
        item["document_code"]
        for item in references
    ]
    retrieved_names = [
        item["document_name"]
        for item in references
    ]

    top1_hit = int(
        bool(retrieved_codes)
        and retrieved_codes[0] == expected_code
    )
    top3_hit = int(
        expected_code in retrieved_codes[:3]
    )
    expected_filename = case.get("expected_filename", "")
    exact_top1_hit = (
        int(
            bool(retrieved_names)
            and retrieved_names[0].casefold()
            == expected_filename.casefold()
        )
        if expected_filename
        else ""
    )
    exact_top3_hit = (
        int(
            expected_filename.casefold()
            in [
                name.casefold()
                for name in retrieved_names[:3]
            ]
        )
        if expected_filename
        else ""
    )

    return {
        **case,
        "status": "success" if result else "failed",
        "attempts": attempts,
        "elapsed_seconds": elapsed,
        "answer": answer,
        "session_id": session_id,
        "actual_top1_document": (
            retrieved_codes[0] if retrieved_codes else ""
        ),
        "actual_top1_filename": (
            retrieved_names[0] if retrieved_names else ""
        ),
        "top1_hit": top1_hit,
        "top3_hit": top3_hit,
        "exact_top1_hit": exact_top1_hit,
        "exact_top3_hit": exact_top3_hit,
        "reference_count": len(references),
        "references": references,
        "error": "" if result else error,
    }


def save_results(records, workbook_path, experiment_label=""):
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = re.sub(
        r"[^A-Za-z0-9_-]+",
        "_",
        experiment_label.strip(),
    ).strip("_")
    label_part = f"_{safe_label}" if safe_label else ""
    base_name = (
        f"batch_eval_{len(records)}"
        f"{label_part}_{timestamp}"
    )

    json_path = RESULTS_DIR / f"{base_name}.json"
    csv_path = RESULTS_DIR / f"{base_name}.csv"

    successful = [
        record
        for record in records
        if record["status"] == "success"
    ]
    exact_scored = [
        record
        for record in successful
        if record.get("expected_filename")
    ]

    summary = {
        "total": len(records),
        "successful": len(successful),
        "failed": len(records) - len(successful),
        "top1_hits": sum(
            record["top1_hit"] for record in successful
        ),
        "top3_hits": sum(
            record["top3_hit"] for record in successful
        ),
        "exact_scored": len(exact_scored),
        "exact_top1_hits": sum(
            record["exact_top1_hit"]
            for record in exact_scored
        ),
        "exact_top3_hits": sum(
            record["exact_top3_hit"]
            for record in exact_scored
        ),
    }

    with json_path.open("w", encoding="utf-8") as file:
        json.dump(
            {
                "generated_at": datetime.now().isoformat(
                    timespec="seconds"
                ),
                "experiment_label": experiment_label,
                "source_workbook": str(workbook_path),
                "summary": summary,
                "results": records,
            },
            file,
            ensure_ascii=False,
            indent=2,
            default=str,
        )

    fieldnames = [
        "question_id",
        "expected_document",
        "expected_filename",
        "question",
        "expected_location",
        "reference_answer",
        "status",
        "attempts",
        "elapsed_seconds",
        "answer",
        "actual_top1_document",
        "actual_top1_filename",
        "top1_hit",
        "top3_hit",
        "exact_top1_hit",
        "exact_top3_hit",
        "reference_count",
        "top3_documents",
        "session_id",
        "error",
    ]

    with csv_path.open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as file:
        writer = csv.DictWriter(
            file,
            fieldnames=fieldnames,
        )
        writer.writeheader()

        for record in records:
            row = {
                key: record.get(key, "")
                for key in fieldnames
            }
            row["top3_documents"] = "|".join(
                reference["document_code"]
                for reference in record["references"][:3]
            )
            writer.writerow(row)

    return json_path, csv_path, summary


def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--limit",
        type=int,
        default=3,
        help="本次运行的题目数量，默认3题",
    )
    parser.add_argument(
        "--label",
        default="",
        help="参数实验标签，将写入结果文件名和JSON",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK_PATH,
        help=(
            "评测工作簿路径；相对路径按项目根目录解析，"
            "默认使用evaluation/baseline中的人工基线工作簿"
        ),
    )
    parser.add_argument(
        "--question-id",
        default="",
        help="只运行指定问题ID，例如H005；默认按工作簿顺序运行",
    )
    parser.add_argument(
        "--question-ids",
        default="",
        help=(
            "运行逗号分隔的多个问题ID，例如E002,E005,E054；"
            "与--question-id不能同时使用"
        ),
    )
    args = parser.parse_args()

    if args.limit < 1:
        print("运行失败：--limit必须大于0")
        return 1

    workbook_path = args.workbook
    if not workbook_path.is_absolute():
        workbook_path = (PROJECT_ROOT / workbook_path).resolve()

    if not workbook_path.is_file():
        print(f"运行失败：找不到评测工作簿：{workbook_path}")
        return 1

    question_id_filter = args.question_id.strip().upper()
    question_id_filters = [
        value.strip().upper()
        for value in args.question_ids.split(",")
        if value.strip()
    ]

    if question_id_filter and question_id_filters:
        print("运行失败：--question-id与--question-ids不能同时使用")
        return 1

    case_limit = (
        len(question_id_filters)
        if question_id_filters
        else args.limit
    )

    try:
        cases = read_test_cases(
            case_limit,
            workbook_path,
            question_id_filter,
            question_id_filters,
        )
        if not cases:
            raise RuntimeError(
                "工作簿中没有找到指定题目"
                if question_id_filter or question_id_filters
                else "工作簿中没有可运行的题目"
            )
        if question_id_filters:
            found_ids = {
                case["question_id"].upper()
                for case in cases
            }
            missing_ids = [
                question_id
                for question_id in question_id_filters
                if question_id not in found_ids
            ]
            if missing_ids:
                raise RuntimeError(
                    "工作簿中没有找到题目："
                    + ", ".join(missing_ids)
                )
        client = RAGFlowClient()
    except Exception as exc:
        print(f"初始化失败：{exc}")
        return 1

    records = []

    for index, case in enumerate(cases, start=1):
        print(
            f"[{index}/{len(cases)}] "
            f"{case['question_id']} "
            f"{case['question']}"
        )

        record = evaluate_case(client, case)
        records.append(record)

        print(
            f"  状态：{record['status']}｜"
            f"尝试：{record['attempts']}｜"
            f"Top1：{record['top1_hit']}｜"
            f"Top3：{record['top3_hit']}｜"
            f"引用：{record['reference_count']}"
        )

    json_path, csv_path, summary = save_results(
        records,
        workbook_path,
        args.label,
    )

    print("=" * 60)
    print(f"总题数：{summary['total']}")
    print(f"成功：{summary['successful']}")
    print(f"失败：{summary['failed']}")
    print(f"Top1命中：{summary['top1_hits']}")
    print(f"Top3命中：{summary['top3_hits']}")
    if summary["exact_scored"]:
        print(
            "精确文件Top1命中："
            f"{summary['exact_top1_hits']}/"
            f"{summary['exact_scored']}"
        )
        print(
            "精确文件Top3命中："
            f"{summary['exact_top3_hits']}/"
            f"{summary['exact_scored']}"
        )
    if args.label:
        print(f"实验标签：{args.label}")
    print(f"评测工作簿：{workbook_path}")
    print(f"JSON结果：{json_path}")
    print(f"CSV结果：{csv_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
